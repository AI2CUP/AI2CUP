"""
ECO (Ethiopian Coffee Organization) Raw Data Processor

Reads weekly ECX export coffee minimum price Excel files from
data/raw/, cleans, normalises, and outputs a simplified dataset.

Output columns: coffee_type, grade, processing, exporter_type,
                year, month, week, price_usd_kg
"""

from __future__ import annotations

import calendar
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ── Unit conversion ──────────────────────────────────────────
LB_TO_KG = 0.45359237

EXCLUDED_COFFEE_TYPES: set[str] = {
    "Roasted Coffee",
    "Coffee Husk and Others",
    "Coffee Husk and Balck Coffee",
    "Coffee Husk and Balck Coffee for extraction",
    "Anaerobic (Preparation)",
    "Carbonic  (Preparation)",
    "Honey (Preparation)",
}


def _usd_lb_to_usd_kg(price_per_lb: float) -> float:
    """Convert USD per pound to USD per kg."""
    return round(price_per_lb / LB_TO_KG, 4)


# ── Date parsing ─────────────────────────────────────────────

_MONTH_NAMES: dict[str, int] = {
    m.lower(): i for i, m in enumerate(calendar.month_name) if m
}
_MONTH_ABBRV: dict[str, int] = {
    m.lower(): i for i, m in enumerate(calendar.month_abbr) if m
}
_ALL_MONTHS = {**_MONTH_NAMES, **_MONTH_ABBRV}

_ETHIOPIAN_MONTHS: dict[str, int] = {
    "meskerem": 1, "tikimt": 2, "hidar": 3, "tahsas": 4,
    "tirr": 5, "yekatit": 6, "megabit": 7, "miazia": 8,
    "ginbot": 9, "sene": 10, "hamle": 11, "nehase": 12, "pagume": 13,
}


def _clamp_date(year: int, month: int, day: int) -> date:
    max_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day, max_day))


def _gregorian_from_ethiopian(month: int, day: int, year_ec: int) -> date:
    year_greg = year_ec + (7 if month <= 4 else 8)
    month_map = {1: 9, 2: 10, 3: 11, 4: 12, 5: 1, 6: 2, 7: 3,
                 8: 4, 9: 5, 10: 6, 11: 7, 12: 8, 13: 9}
    return _clamp_date(year_greg, month_map.get(month, 1), day)


def _normalise_month_name(name: str) -> str:
    name = name.strip().rstrip(".").lower()
    mapping = {
        "jan": "january", "feb": "february", "mar": "march",
        "apr": "april", "may": "may", "jun": "june", "jul": "july",
        "aug": "august", "sep": "september", "oct": "october",
        "nov": "november", "dec": "december", "octo": "october",
        "nove": "november",
    }
    return mapping.get(name, name)


def _parse_week_start(text: str) -> date | None:
    """Parse the week-start date from an ECO filename or title."""
    cl = text.upper()
    clean = re.sub(r"\(?(NEW|OLD)\s*CROP\s*[^)]*\)?", "", text, flags=re.IGNORECASE)
    clean = re.sub(r"Export Coffee Minimum Price Data", "", clean, flags=re.IGNORECASE)
    clean = re.sub(r"[()]", "", clean)

    # Ethiopian calendar
    m = re.search(
        rf"({'|'.join(_ETHIOPIAN_MONTHS)})\s+(\d+)[/\s-]*(\d{{4}})",
        clean, re.IGNORECASE,
    )
    if m:
        return _gregorian_from_ethiopian(
            _ETHIOPIAN_MONTHS[m.group(1).lower()],
            int(m.group(2)), int(m.group(3)),
        )

    # Cross-month / standard range / "to" variant
    patterns = [
        r"(\w+\.?)\s+(\d+)\s*[-–—to]+\s*(\w+\.?)\s+(\d+),?\s*(\d{4})",
        r"(\w+\.?)\s+(\d+)\s*[-–—]+\s*(\d+),?\s*(\d{4})",
        r"(\w+\.?)\s+(\d+)\s+to\s+(\w+\.?)\s+(\d+),?\s*(\d{4})",
    ]
    for pat in patterns:
        m = re.search(pat, clean, re.IGNORECASE)
        if m:
            groups = m.groups()
            month_name = _normalise_month_name(groups[0])
            start_day = int(groups[1])
            year = int(groups[-1])
            month_num = _ALL_MONTHS.get(month_name)
            if month_num:
                return _clamp_date(year, month_num, start_day)

    # Underscore separated: Feb_18_24_2025
    m = re.search(r"(\w+)_(\d+)_(\d+)_(\d{4})", text, re.IGNORECASE)
    if m:
        month_num = _ALL_MONTHS.get(_normalise_month_name(m.group(1)))
        if month_num:
            return _clamp_date(int(m.group(4)), month_num, int(m.group(2)))

    # Dot variant: Dec 31.2024
    m = re.search(r"(\w+\.?)\s+(\d+)\.(\d{4})", clean, re.IGNORECASE)
    if m:
        month_num = _ALL_MONTHS.get(_normalise_month_name(m.group(1)))
        if month_num:
            return _clamp_date(int(m.group(3)), month_num, int(m.group(2)))

    # Fallback: any month + 4-digit year
    m = re.search(r"(\w+)\s+(\d{4})", clean)
    if m:
        month_num = _ALL_MONTHS.get(_normalise_month_name(m.group(1)))
        if month_num:
            return date(int(m.group(2)), month_num, 15)

    return None


# ── Data cleaning ────────────────────────────────────────────

def _clean_grade(raw: Any) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw) if 1 <= int(raw) <= 5 else None
    s = str(raw).strip().upper()
    if s in ("UG", "PEABER", "PEABERRY"):
        return None
    try:
        v = int(s)
        return v if 1 <= v <= 5 else None
    except (ValueError, TypeError):
        return None


def _clean_coffee_type(raw: str) -> str | None:
    ct = str(raw).strip()
    if not ct or ct in EXCLUDED_COFFEE_TYPES:
        return None
    # Normalise common variants
    mapping = {
        "Andrecha": "Andracha",
        "Djimma": "Jimma",
        "Limmu": "Limu",
        "sheka": "Sheka",
        "kafa": "Kafa",
        "bonga": "Bonga",
        "Bench maji": "Bench Maji",
        "Galana Abya": "Galana Abaya",
        "Gelana Abya": "Galana Abaya",
        "Mokasida": "Mokamba",
        "Goma": "Gera",
    }
    return mapping.get(ct, ct)


def _normalise_processing(raw: str) -> str:
    r = raw.strip().lower()
    if r in ("natural", "unwashed"):
        return "Natural"
    if r in ("washed",):
        return "Washed"
    return ""


def _normalise_exporter(raw: str) -> str:
    r = str(raw).strip().lower()
    if r == "commercial":
        return "Commercial"
    if r == "grower":
        return "Grower"
    if r == "union":
        return "Union"
    if r in ("v/integration", "vintegration"):
        return "V/Integration"
    return "Commercial"


# ── Main processor ───────────────────────────────────────────

RAW_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "raw"


def process_all_raw_files(raw_dir: str | Path = RAW_DIR) -> pd.DataFrame:
    """Read all ECO Excel files, returning a simplified DataFrame."""
    raw_dir = Path(raw_dir)
    records: list[dict] = []

    for fpath in sorted(raw_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            continue

        if ws.max_column is not None and ws.max_column < 7:
            wb.close()
            continue

        # Parse date from title cell B1 or filename
        week_start = None
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            if row and len(row) > 1 and row[1] and isinstance(row[1], str) and len(row[1]) > 10:
                week_start = _parse_week_start(row[1])
                break
        if week_start is None:
            week_start = _parse_week_start(fpath.stem)
        if week_start is None:
            wb.close()
            continue

        year = week_start.year
        month = week_start.month
        week = week_start.isocalendar().week

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 7:
                continue
            coffee_type = _clean_coffee_type(str(row[2]).strip() if row[2] else "")
            if coffee_type is None:
                continue
            grade = _clean_grade(row[4])
            if grade is None:
                continue
            processing = _normalise_processing(str(row[5]) if row[5] else "")
            if not processing:
                continue
            exporter_type = _normalise_exporter(str(row[3]) if row[3] else "")
            try:
                price_usd_lb = float(row[6])
            except (ValueError, TypeError):
                continue
            if price_usd_lb <= 0:
                continue

            records.append({
                "coffee_type": coffee_type,
                "grade": grade,
                "processing": processing,
                "exporter_type": exporter_type,
                "year": year,
                "month": month,
                "week": int(week),
                "price_usd_kg": _usd_lb_to_usd_kg(price_usd_lb),
            })

        wb.close()

    df = pd.DataFrame(records)
    # Deduplicate
    df = df.drop_duplicates()
    return df


if __name__ == "__main__":
    print("ECO Data Processor — processing raw files...")
    df = process_all_raw_files()
    out_path = (
        Path(__file__).resolve().parent.parent.parent
        / "data" / "storage" / "eco_training.parquet"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(out_path, index=False)
    df.to_csv(out_path.with_suffix(".csv"), index=False)
    print(f"\nProcessed {len(df)} records -> {out_path}")
    print(f"\nColumns: {list(df.columns)}")
    print(f"\nPrice stats (USD/kg):")
    print(f"  Min: {df['price_usd_kg'].min():.2f}")
    print(f"  Max: {df['price_usd_kg'].max():.2f}")
    print(f"  Mean: {df['price_usd_kg'].mean():.2f}")
    print(f"  Std: {df['price_usd_kg'].std():.2f}")
    print(f"\nRecords per processing:")
    print(df["processing"].value_counts().to_string())
    print(f"\nRecords per grade:")
    print(df["grade"].value_counts().sort_index().to_string())
